import openpyxl
import pandas as pd
import json
from datetime import datetime, timedelta
import os

import unidecode

class HourData:
    def __init__(self, zones_array):
        self.zones_array = zones_array
    
    def __str__(self):
        return f"Hour Data:\nZones: {self.zones_array}"
    def to_dict(self):
         
             array = [zone.to_dict() for zone in self.zones_array]
             return {
                 "aforo_clases": array
             }

class Zones_Data:
    def __init__(self, name, targetCapacity, totalCapacity,zone,hour):
        self.name = name
        self.targetCapacity = targetCapacity
        self.totalCapacity = totalCapacity
        self.zone = zone
        self.hour = hour
    
    def __str__(self):
        return f"Zones Data:\nName: {self.name}\nTarget Capacity: {self.targetCapacity}\nTotal Capacity: {self.totalCapacity} \nZone : {self.zone}"
    def to_dict(self):
        return {
            "name": self.name,
            "zone":self.zone,
            "targetCapacity": self.targetCapacity,
            "totalCapacity": self.totalCapacity,
            "hour":self.hour
            
        }
def serialize_to_json(obj, filename):
    # Initialize data from the new object
    new_data = obj.to_dict()
    
    # If file exists, read and append to existing data
    if os.path.exists(filename):
        with open(filename, 'r') as f:
            existing_data = json.load(f)
            
            # Create a set of existing class names to check for duplicates
            existing_names = {item['name'] for item in existing_data["aforo_clases"]}
            
            # Only append classes that aren't already in the file
            new_classes = [cls for cls in new_data["aforo_clases"] 
                         if cls['name'] not in existing_names]
            
            existing_data["aforo_clases"].extend(new_classes)
            final_data = existing_data
    else:
        # If file doesn't exist, use new data directly
        final_data = new_data
    
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    with open(filename, 'w') as f:
        json.dump(final_data, f, indent=4)
    
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    with open(filename, 'w') as f:
        json.dump(final_data, f, indent=4)
def clean_string(text):
    if(text!=None):
        text_without_accents = unidecode.unidecode(text)
        cleaned_text = ''.join(char for char in text_without_accents if ord(char) < 128)
        return cleaned_text
    return None
def get_cell_merge_info(sheet,row, column):
    # Convertir la columna a número si se proporciona como letra
    if isinstance(column, str):
        column = openpyxl.utils.column_index_from_string(column)

    # Verificar si la celda está en un rango fusionado
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row <= row <= max_row and min_col <= column <= max_col:
            num_cols = max_col - min_col + 1
            num_rows = max_row - min_row + 1
            return {
                "is_merged": True,
                "columns_span": num_cols,
                "rows_span": num_rows,
                "value": sheet.cell(row=min_row, column=min_col).value
            }

    # Si la celda no está fusionada
    return {
        "is_merged": False,
        "columns_span": 1,
        "rows_span": 1,
        "value": sheet.cell(row=row, column=column).value
    }
def modify_event(string):
    # Map the activities to their new names
    activity_map = {
        "CICLO INDOOR": "Studio 3",
        "FUNCIONAL 360": "FUNCIONAL",
        "PILATES": "Studio 2",
        "BODY COMBAT": "Studio 1",
        "AQUAGYM": "PP",
        "ZUMBA": "Studio 4",
        "BODY STEP": "Studio 1",
        "DEPORTE I (3-7)": "Studio 4",
        "BODY BALANCE":"Studio 4",
        "BODY STEP":"Studio 1",
        "X-TRAINING KIDS":"X-TRAINING",
        "GAP":"Studio 2",
        "X-TRAINING":"X-TRAINING",
        "BODY PUMP":"Studio 1",
        "AQUAFIT":"PP",
        "EMBARAZO ACT.":"Studio 2",
        "SUELO PÉLVICO":"Studio 2",
        "DANCE":"Studio 2",
        "ZUMBA KIDS":"X-TRAINING",
        "EN FAMILIA":"Studio 2",
    }
    return  activity_map.get(string, "Studio 1")
    

def redondear_hora(hora_str):
    # Round the time to the nearest half hour
    hora = datetime.strptime(hora_str, "%H:%M")
    minuto = hora.minute
    if minuto >= 45:
        hora += timedelta(hours=1)
        hora = hora.replace(minute=0)
    elif minuto >= 15:
        hora = hora.replace(minute=30)
    else:
        hora = hora.replace(minute=0)
    return hora.strftime("%H")

def generar_json_por_hora(file_path,startDate,output_file):
    from datetime import datetime, timedelta

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    day = startDate
    
    
    # Create base date starting from September 1st, 2024
    base_date = datetime(2024, 9, 2)
    

    for col in range(3, 28, 4):
        rowStart = 3
        current_date = base_date + timedelta(days=day-1)  # subtract 1 since startDate begins at 1
        formatted_date = current_date.strftime('%Y-%m-%d')
        while(rowStart < 94-1):
            notWorkingHour = False
            zones_data = []
            hour = sheet.cell(row=rowStart, column=1).value
            
            value = get_cell_merge_info(sheet, rowStart, 1)
            sumatorio = value['rows_span']            
            rowend = rowStart + sumatorio
            for row in range(rowStart, rowend, 2):
                name = sheet.cell(row=row, column=col-1).value
                
                if name == None or name == "":
                    break;
                cleanName = clean_string(name)
                total = sheet.cell(row=row+1, column=col).value
                target = sheet.cell(row=row+1, column=col+1).value
                
                data = Zones_Data(cleanName, target, total,modify_event(cleanName),hour)
                
                zones_data.append(data)
                
            
            rowStart += sumatorio
            value = HourData(zones_data)
            if(len(value.zones_array)!=0):

                serialize_to_json(value, f"{output_file}{formatted_date}_{hour[:-3]}.json")

        day = day + 1        
        


def generate_clases():
    startDate = 1
    output_file = "data/formated_data/zones/"
    paths=["data/excel/clases_sept_2-8.xlsx",
           "data/excel/clases_sept_9-15.xlsx",
           "data/excel/clases_sept_16-22.xlsx",
           "data/excel/clases_sept_23-29.xlsx"]
    for path in paths:
        generar_json_por_hora(path,startDate,output_file)
        startDate+=7
    output_file = "data/formated_data/averageZones/"
    startDate = 1
    generar_json_por_hora("data/excel/avg-clases-sept.xlsx",startDate,output_file)

