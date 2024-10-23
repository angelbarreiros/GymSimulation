import math
import openpyxl
import json

import unidecode


class Day_Data:
    def __init__(self, name, hours_array):
        self.name = name
        self.hours_array = hours_array
    
    def __str__(self):
        return f"Day Data:\nName: {self.name}\nHours: {self.hours_array}"
    def to_dict(self):
        return {
            "name": self.name,
            "hours_array": [hour.to_dict() for hour in self.hours_array]
        }
   

class HourData:
    def __init__(self, zones_array):
        self.zones_array = zones_array
    
    def __str__(self):
        return f"Hour Data:\nZones: {self.zones_array}"
    def to_dict(self):
         
             array = [zone.to_dict() for zone in self.zones_array]
             return {
                 "aforo_zonas": array
             }
        

class Zones_Data:
    def __init__(self, name, targetCapacity, totalCapacity):
        self.name = name
        self.targetCapacity = targetCapacity
        self.totalCapacity = totalCapacity
    
    def __str__(self):
        return f"Zones Data:\nName: {self.name}\nTarget Capacity: {self.targetCapacity}\nTotal Capacity: {self.totalCapacity}"
    def to_dict(self):
        return {
            "name": self.name,
            "targetCapacity": self.targetCapacity,
            "totalCapacity": self.totalCapacity
        }
    1
def serialize_to_json(obj, filename):
    with open(filename, 'w') as f:
        json.dump(obj.to_dict(), f, indent=4)

def clean_string(text):
    if(text!=None):
        text_without_accents = unidecode.unidecode(text)
        cleaned_text = ''.join(char for char in text_without_accents if ord(char) < 128)
        return cleaned_text
    return None
def get_cell_merge_info(sheet,row, column):
    # Convertir la columna a número si se proporciona como letra
    if isinstance(column, str):
        column = openpyxl.utils.column_index_from_string(column)

    # Verificar si la celda está en un rango fusionado
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row <= row <= max_row and min_col <= column <= max_col:
            num_cols = max_col - min_col + 1
            num_rows = max_row - min_row + 1
            return {
                "is_merged": True,
                "columns_span": num_cols,
                "rows_span": num_rows,
                "value": sheet.cell(row=min_row, column=min_col).value
            }

    # Si la celda no está fusionada
    return {
        "is_merged": False,
        "columns_span": 1,
        "rows_span": 1,
        "value": sheet.cell(row=row, column=column).value
    }

def excel_to_json(file_path, startDate):
    from datetime import datetime, timedelta

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    day = startDate
    
    # Create base date starting from September 1st, 2024
    base_date = datetime(2024, 9, 2)
    

    for col in range(3, 22, 3):
        rowStart = 3
        rowend = 31
        current_date = base_date + timedelta(days=day-1)  # subtract 1 since startDate begins at 1
        formatted_date = current_date.strftime('%Y-%m-%d')
        while(rowStart < sheet.max_row -1):
            notWorkingHour = False
            zones_data = []
            hour = sheet.cell(row=rowStart, column=1).value
            value = get_cell_merge_info(sheet, rowStart, 1)
            sumatorio = value['rows_span']
            
            # Format current date
            
            
            for row in range(rowStart, rowend, 2):
                name = sheet.cell(row=row, column=col-1).value
                if name == None or name == "":
                    notWorkingHour = True
                cleanName = clean_string(name)
                total = sheet.cell(row=row+1, column=col).value
                target = sheet.cell(row=row+1, column=col+1).value
                data = Zones_Data(cleanName, target, total)
                zones_data.append(data)
                
            rowend += sumatorio
            rowStart += sumatorio
            value = HourData(zones_data)
            
            if not notWorkingHour:
                serialize_to_json(value, f"data/formated_data/zones/{formatted_date}_{hour[:-3]}.json")
            notWorkingHour = False
        day = day + 1        
        

def getZonesData():  
    startDate = 1
    paths=["data/excel/zones-sept-1-8.xlsx",
           "data/excel/zones-sept-8-15.xlsx",
           "data/excel/zones-sept-15-22.xlsx",
           "data/excel/zones-sept-22-30.xlsx"]
    for path in paths:
        excel_to_json(path,startDate= startDate)
        startDate+=7